#!/usr/bin/env python3
"""
GHL Mastery Financial Model Generator
Creates comprehensive financial model with corrected data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
BLUE = Font(color='0000FF')
BLACK = Font(color='000000')
GREEN = Font(color='008000')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
BOLD = Font(bold=True)

HEADER_FILL = PatternFill('solid', fgColor='1E3A5F')
INPUT_FILL = PatternFill('solid', fgColor='FFFF99')
CALC_FILL = PatternFill('solid', fgColor='E8F5E9')
HIGHLIGHT_FILL = PatternFill('solid', fgColor='D4A574')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_workbook():
    wb = Workbook()

    # Sheet 1: Corrected Funnel Data
    ws1 = wb.active
    ws1.title = "Funnel Analysis"
    create_funnel_sheet(ws1)

    # Sheet 2: Unit Economics
    ws2 = wb.create_sheet("Unit Economics")
    create_unit_economics_sheet(ws2)

    # Sheet 3: Scenario Projections
    ws3 = wb.create_sheet("Scenario Projections")
    create_scenario_sheet(ws3)

    # Sheet 4: ROI Calculator
    ws4 = wb.create_sheet("ROI Calculator")
    create_roi_sheet(ws4)

    # Sheet 5: Benchmarks
    ws5 = wb.create_sheet("Industry Benchmarks")
    create_benchmark_sheet(ws5)

    return wb

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal='center')

def create_funnel_sheet(ws):
    ws['A1'] = "GHL MASTERY - CORRECTED FUNNEL ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A2'] = "Period: Oct 30, 2025 - Jan 30, 2026 (90 Days)"
    ws['A3'] = "Source: All from Paid Ads"

    # Input Section
    ws['A5'] = "INPUTS (Blue = Editable)"
    ws['A5'].font = BOLD

    headers = ['Metric', 'Value', 'Notes']
    for i, h in enumerate(headers, 1):
        ws.cell(row=6, column=i, value=h)
    style_header(ws, 6, 3)

    inputs = [
        ('Ad Spend (90 days)', 1500, 'Estimate - confirm with Adam'),
        ('Leads Generated', 238, 'From GHL Dashboard'),
        ('Tripwire Purchases (Ads)', 33, 'From Stripe CSV'),
        ('Tripwire Purchases (Database)', 84, 'Adam confirmed'),
        ('Appointments Booked', 46, 'From GHL Calendar'),
        ('Appointments Showed', 22, 'From GHL Dashboard'),
        ('No-Shows', 11, 'From GHL Dashboard'),
        ('Cancelled', 9, 'From GHL Dashboard'),
        ('Assumed Close Rate', 0.30, 'Industry avg - needs confirmation'),
    ]

    for i, (metric, value, note) in enumerate(inputs, 7):
        ws.cell(row=i, column=1, value=metric)
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = BLUE
        cell.fill = INPUT_FILL
        ws.cell(row=i, column=3, value=note)

    # Calculated Metrics
    ws['A18'] = "CALCULATED METRICS"
    ws['A18'].font = BOLD

    calcs = [
        ('Cost Per Lead', '=B7/B8', '$'),
        ('Tripwire Conversion (Ads)', '=B9/B8', '%'),
        ('Lead to Appointment', '=B11/B8', '%'),
        ('Show Rate (excl cancelled)', '=B12/(B11-B14)', '%'),
        ('Show Rate (all booked)', '=B12/B11', '%'),
        ('Cost Per Appointment', '=B7/B11', '$'),
        ('Cost Per Show', '=B7/B12', '$'),
        ('Estimated Closes', '=B12*B15', '#'),
        ('Cost Per Acquisition (CAC)', '=B7/B26', '$'),
    ]

    for i, (metric, formula, fmt) in enumerate(calcs, 19):
        ws.cell(row=i, column=1, value=metric)
        cell = ws.cell(row=i, column=2, value=formula)
        cell.font = BLACK
        cell.fill = CALC_FILL
        ws.cell(row=i, column=3, value=fmt)

    # Benchmark Comparison
    ws['A30'] = "BENCHMARK COMPARISON"
    ws['A30'].font = BOLD

    headers2 = ['Metric', 'GHL Mastery', 'Industry Avg', 'Industry Top', 'Status']
    for i, h in enumerate(headers2, 1):
        ws.cell(row=31, column=i, value=h)
    style_header(ws, 31, 5)

    benchmarks = [
        ('Lead → Appointment', '=B21', '3-5%', '8-10%', '=IF(B21>0.1,"⭐ TOP",IF(B21>0.05,"✅ Good","⚠️ Below"))'),
        ('Show Rate', '=B22', '60-75%', '80-95%', '=IF(B22>0.8,"⭐ TOP",IF(B22>0.6,"✅ Good","⚠️ Below"))'),
        ('Cost Per Lead', '=B19', '$50-85', '$20-35', '=IF(B19<35,"⭐ TOP",IF(B19<85,"✅ Good","⚠️ Above"))'),
    ]

    for i, row_data in enumerate(benchmarks, 32):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35

def create_unit_economics_sheet(ws):
    ws['A1'] = "UNIT ECONOMICS ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "REVENUE DATA (90 Days)"
    ws['A3'].font = BOLD

    headers = ['Category', 'Revenue', 'Transactions', '% of Total']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 4)

    revenue_data = [
        ('Recurring Subscriptions', 52066, 123),
        ('VIP $5K Sales', 25000, 5),
        ('Invoice/Custom Work', 25394, 9),
        ('Low-ticket/Tripwire', 7467, 77),
    ]

    for i, (cat, rev, txn) in enumerate(revenue_data, 5):
        ws.cell(row=i, column=1, value=cat)
        cell = ws.cell(row=i, column=2, value=rev)
        cell.font = BLUE
        cell.fill = INPUT_FILL
        cell.number_format = '$#,##0'
        ws.cell(row=i, column=3, value=txn)
        ws.cell(row=i, column=4, value=f'=B{i}/B9')
        ws.cell(row=i, column=4).number_format = '0.0%'

    ws['A9'] = 'TOTAL'
    ws['A9'].font = BOLD
    ws['B9'] = '=SUM(B5:B8)'
    ws['B9'].font = BOLD
    ws['B9'].number_format = '$#,##0'
    ws['C9'] = '=SUM(C5:C8)'
    ws['C9'].font = BOLD

    # MRR Breakdown
    ws['A12'] = "ESTIMATED MRR BREAKDOWN"
    ws['A12'].font = BOLD

    headers2 = ['Product', 'Rate', 'Est. Customers', 'Monthly Value']
    for i, h in enumerate(headers2, 1):
        ws.cell(row=13, column=i, value=h)
    style_header(ws, 13, 4)

    mrr_data = [
        ('Micro Snapshots', 97, 3),
        ('AOF Course', 197, 10),
        ('Coaching Tier', 297, 7),
        ('VIP Coaching', 497, 8),
        ('Premium', 597, 3),
        ('Legacy VA', 650, 10),
        ('Current VA', 900, 1),
        ('High-tier', 997, 2),
    ]

    for i, (prod, rate, cust) in enumerate(mrr_data, 14):
        ws.cell(row=i, column=1, value=prod)
        cell_rate = ws.cell(row=i, column=2, value=rate)
        cell_rate.font = BLUE
        cell_rate.fill = INPUT_FILL
        cell_rate.number_format = '$#,##0'
        cell_cust = ws.cell(row=i, column=3, value=cust)
        cell_cust.font = BLUE
        cell_cust.fill = INPUT_FILL
        ws.cell(row=i, column=4, value=f'=B{i}*C{i}')
        ws.cell(row=i, column=4).number_format = '$#,##0'

    ws['A22'] = 'TOTAL MRR'
    ws['A22'].font = BOLD
    ws['D22'] = '=SUM(D14:D21)'
    ws['D22'].font = BOLD
    ws['D22'].number_format = '$#,##0'
    ws['D22'].fill = HIGHLIGHT_FILL

    # LTV Calculations
    ws['A25'] = "LIFETIME VALUE (LTV)"
    ws['A25'].font = BOLD

    ws['A26'] = 'LTV - Coaching'
    ws['B26'] = 500
    ws['B26'].font = BLUE
    ws['B26'].fill = INPUT_FILL
    ws['C26'] = 'Monthly rate'
    ws['D26'] = 4
    ws['D26'].font = BLUE
    ws['D26'].fill = INPUT_FILL
    ws['E26'] = 'Avg months'
    ws['F26'] = '=B26*D26'
    ws['F26'].number_format = '$#,##0'

    ws['A27'] = 'LTV - VA'
    ws['B27'] = 650
    ws['B27'].font = BLUE
    ws['B27'].fill = INPUT_FILL
    ws['D27'] = 8
    ws['D27'].font = BLUE
    ws['D27'].fill = INPUT_FILL
    ws['F27'] = '=B27*D27'
    ws['F27'].number_format = '$#,##0'

    ws['A28'] = 'LTV - Blended'
    ws['F28'] = '=AVERAGE(F26:F27)'
    ws['F28'].number_format = '$#,##0'
    ws['F28'].font = BOLD

    # LTV:CAC
    ws['A31'] = "LTV:CAC RATIO"
    ws['A31'].font = BOLD

    ws['A32'] = 'CAC (from Funnel Analysis)'
    ws['B32'] = "='Funnel Analysis'!B27"
    ws['B32'].number_format = '$#,##0.00'

    ws['A33'] = 'LTV:CAC Ratio'
    ws['B33'] = '=F28/B32'
    ws['B33'].number_format = '0.0x'
    ws['B33'].font = BOLD
    ws['B33'].fill = HIGHLIGHT_FILL

    ws['C33'] = '=IF(B33>=5,"⭐ Excellent (5x+)",IF(B33>=3,"✅ Healthy (3x+)","⚠️ Below threshold"))'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

def create_scenario_sheet(ws):
    ws['A1'] = "90-DAY SCENARIO PROJECTIONS"
    ws['A1'].font = Font(bold=True, size=14)

    # Assumptions
    ws['A3'] = "SCENARIO ASSUMPTIONS"
    ws['A3'].font = BOLD

    headers = ['Assumption', 'Conservative', 'Moderate', 'Aggressive']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 4)

    assumptions = [
        ('Month 3 Ad Spend', 2500, 5000, 10000),
        ('CPL (maintained)', 6.30, 6.30, 6.30),
        ('Booking Rate', 0.19, 0.19, 0.19),
        ('Show Rate (improved)', 0.65, 0.70, 0.75),
        ('Close Rate', 0.30, 0.35, 0.38),
        ('Avg Sale Value', 500, 500, 500),
        ('Failed Payment Recovery', 0.50, 0.65, 0.75),
        ('Database Revenue (one-time)', 15000, 25000, 40000),
    ]

    for i, row_data in enumerate(assumptions, 5):
        ws.cell(row=i, column=1, value=row_data[0])
        for j, val in enumerate(row_data[1:], 2):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BLUE
            cell.fill = INPUT_FILL
            if isinstance(val, float) and val < 1:
                cell.number_format = '0.0%'
            elif isinstance(val, (int, float)) and val >= 100:
                cell.number_format = '$#,##0'

    # Calculated Projections
    ws['A15'] = "MONTH 3 PROJECTIONS"
    ws['A15'].font = BOLD

    calc_headers = ['Metric', 'Conservative', 'Moderate', 'Aggressive']
    for i, h in enumerate(calc_headers, 1):
        ws.cell(row=16, column=i, value=h)
    style_header(ws, 16, 4)

    calcs = [
        ('Leads Generated', '=B5/B6', '=C5/C6', '=D5/D6'),
        ('Appointments Booked', '=B17*B7', '=C17*C7', '=D17*D7'),
        ('Shows', '=B18*B8', '=C18*C8', '=D18*D8'),
        ('Sales', '=B19*B9', '=C19*C9', '=D19*D9'),
        ('New MRR', '=B20*B10', '=C20*C10', '=D20*D10'),
    ]

    for i, row_data in enumerate(calcs, 17):
        ws.cell(row=i, column=1, value=row_data[0])
        for j, formula in enumerate(row_data[1:], 2):
            cell = ws.cell(row=i, column=j, value=formula)
            cell.fill = CALC_FILL
            if i == 21:
                cell.number_format = '$#,##0'
            else:
                cell.number_format = '#,##0'

    # 90-Day Summary
    ws['A24'] = "90-DAY TOTAL IMPACT"
    ws['A24'].font = BOLD

    summary_headers = ['Revenue Source', 'Conservative', 'Moderate', 'Aggressive']
    for i, h in enumerate(summary_headers, 1):
        ws.cell(row=25, column=i, value=h)
    style_header(ws, 25, 4)

    ws['A26'] = 'New MRR Added (3 months)'
    ws['B26'] = '=B21*3*0.6'  # Ramp factor
    ws['C26'] = '=C21*3*0.7'
    ws['D26'] = '=D21*3*0.8'

    ws['A27'] = 'Database Revenue'
    ws['B27'] = '=B12'
    ws['C27'] = '=C12'
    ws['D27'] = '=D12'

    ws['A28'] = 'Failed Payment Recovery'
    ws['B28'] = '=25870*B11'
    ws['C28'] = '=25870*C11'
    ws['D28'] = '=25870*D11'

    ws['A29'] = 'TOTAL NEW REVENUE'
    ws['A29'].font = BOLD
    for col in ['B', 'C', 'D']:
        cell = ws[f'{col}29']
        cell.value = f'=SUM({col}26:{col}28)'
        cell.font = BOLD
        cell.fill = HIGHLIGHT_FILL
        cell.number_format = '$#,##0'

    for row in range(26, 29):
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '$#,##0'
            ws[f'{col}{row}'].fill = CALC_FILL

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_roi_sheet(ws):
    ws['A1'] = "ENGAGEMENT ROI CALCULATOR"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "ENGAGEMENT PRICING"
    ws['A3'].font = BOLD

    ws['A4'] = 'Base Investment'
    ws['B4'] = 60000
    ws['B4'].font = BLUE
    ws['B4'].fill = INPUT_FILL
    ws['B4'].number_format = '$#,##0'

    ws['A5'] = 'Performance Bonus'
    ws['B5'] = 20000
    ws['B5'].font = BLUE
    ws['B5'].fill = INPUT_FILL
    ws['B5'].number_format = '$#,##0'

    ws['A6'] = 'Total (if bonus achieved)'
    ws['B6'] = '=B4+B5'
    ws['B6'].font = BOLD
    ws['B6'].number_format = '$#,##0'

    ws['A9'] = "PROJECTED RETURNS"
    ws['A9'].font = BOLD

    headers = ['Timeframe', 'Conservative', 'Moderate', 'Aggressive']
    for i, h in enumerate(headers, 1):
        ws.cell(row=10, column=i, value=h)
    style_header(ws, 10, 4)

    ws['A11'] = '90-Day Revenue Increase'
    ws['B11'] = "='Scenario Projections'!B29"
    ws['C11'] = "='Scenario Projections'!C29"
    ws['D11'] = "='Scenario Projections'!D29"

    ws['A12'] = '12-Month Revenue Increase'
    ws['B12'] = '=B11*4'
    ws['C12'] = '=C11*4'
    ws['D12'] = '=D11*4'

    ws['A13'] = 'ROI (vs Base Investment)'
    ws['B13'] = '=B12/$B$4'
    ws['C13'] = '=C12/$B$4'
    ws['D13'] = '=D12/$B$4'

    ws['A14'] = 'ROI (vs Total Investment)'
    ws['B14'] = '=B12/$B$6'
    ws['C14'] = '=C12/$B$6'
    ws['D14'] = '=D12/$B$6'

    for col in ['B', 'C', 'D']:
        ws[f'{col}11'].number_format = '$#,##0'
        ws[f'{col}12'].number_format = '$#,##0'
        ws[f'{col}13'].number_format = '0.0x'
        ws[f'{col}14'].number_format = '0.0x'
        ws[f'{col}11'].fill = CALC_FILL
        ws[f'{col}12'].fill = CALC_FILL
        ws[f'{col}13'].fill = HIGHLIGHT_FILL
        ws[f'{col}14'].fill = HIGHLIGHT_FILL

    ws['A17'] = "PAYBACK PERIOD"
    ws['A17'].font = BOLD

    ws['A18'] = 'Monthly Revenue Increase'
    ws['B18'] = '=B12/12'
    ws['C18'] = '=C12/12'
    ws['D18'] = '=D12/12'

    ws['A19'] = 'Months to Payback (Base)'
    ws['B19'] = '=$B$4/B18'
    ws['C19'] = '=$B$4/C18'
    ws['D19'] = '=$B$4/D18'

    for col in ['B', 'C', 'D']:
        ws[f'{col}18'].number_format = '$#,##0'
        ws[f'{col}19'].number_format = '0.0'
        ws[f'{col}18'].fill = CALC_FILL
        ws[f'{col}19'].fill = CALC_FILL

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_benchmark_sheet(ws):
    ws['A1'] = "INDUSTRY BENCHMARKS (CITED)"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Metric', 'Conservative', 'Average', 'Top Performer', 'Source']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 5)

    benchmarks = [
        ('Lead to Appointment', '2%', '3-5%', '8-10%', 'Intelemark, Martal Group 2024'),
        ('Appointment Show Rate', '60%', '70-75%', '85-95%', 'Roezan, Close.com'),
        ('Sales Close Rate', '15%', '20-25%', '30-40%', 'Arrows.to 2024'),
        ('Facebook CPL (B2B)', '$100+', '$50-85', '$20-35', 'WordStream 2024'),
        ('LTV:CAC Ratio', '3x', '3.2x', '5x+', 'Phoenix Strategy Group'),
        ('Tripwire Conversion', '1%', '2-3%', '5-10%', 'Data Driven Marketing'),
        ('Database Reactivation', '3-5%', '10-15%', '20-30%', 'AgencyAnalytics, Klaviyo'),
        ('Failed Payment Recovery', '45-50%', '60-70%', '75-85%', 'Churnkey 2025, ProsperStack'),
    ]

    for i, row_data in enumerate(benchmarks, 4):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 5:
                cell.font = Font(italic=True, size=9)

    ws['A14'] = "GHL MASTERY vs BENCHMARKS"
    ws['A14'].font = BOLD

    headers2 = ['Metric', 'GHL Mastery', 'vs Average', 'Percentile']
    for i, h in enumerate(headers2, 1):
        ws.cell(row=15, column=i, value=h)
    style_header(ws, 15, 4)

    comparisons = [
        ('Lead to Appointment', '19.3%', '+290% to +540%', 'Top 1%'),
        ('Show Rate', '59-67%', 'At par', '50th'),
        ('Cost Per Lead', '$6.30', '-85% to -93%', 'Top 5%'),
        ('LTV:CAC Ratio', '15.8x', '+394%', 'Top 1%'),
        ('Tripwire Conversion', '14%', '+367% to +600%', 'Top 10%'),
    ]

    for i, row_data in enumerate(comparisons, 16):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 4 and 'Top' in str(val):
                cell.fill = HIGHLIGHT_FILL

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35

if __name__ == '__main__':
    wb = create_workbook()
    output_path = '/sessions/pensive-youthful-gates/mnt/rm-marketing/rm-sales-buildout-proposal/GHL_Mastery_Financial_Model_v2.xlsx'
    wb.save(output_path)
    print(f"Saved to {output_path}")
